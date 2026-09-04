"""Tests des générateurs de rapports (XLSX, DOCX, PDF) et des graphiques."""

import os
import re
import shutil
import struct
import sys
import tempfile
import unittest
import zipfile
import zlib
import xml.etree.ElementTree as ET

sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "src"))

from bassin.core import rainfall  # noqa: E402
from bassin.core.model import Bassin, BassinAmont, Projet  # noqa: E402
from bassin.reports import charts, docx_report, dossier as mod_dossier, pdf_report, xlsx_report  # noqa: E402
from bassin.reports.pdf_writer import Pdf, largeur_texte, nettoyer  # noqa: E402


def projet_complet() -> Projet:
    p = Projet(commune_ins="63013", commune_nom="Bütgenbach", periode_retour=25,
               surfaces=Projet.surfaces_par_defaut(), surface_reference_m2=2000.0,
               nom_projet="Lotissement Les Sources", auteur="Bureau d'études", localisation="Rue du Moulin")
    p.surfaces[7].aire_m2 = 1500.0
    p.surfaces[1].aire_m2 = 500.0
    p.k_infiltration_ms = 1e-5
    p.surface_infiltration_m2 = 120.0
    p.debit_ajutage_ls = 1.0
    p.bassin = Bassin(volume_total_m3=90.0, volume_sous_ajutage_m3=10.0,
                      surface_dispersion_m2=120.0, debit_ajutage_ls=1.0)
    return p


class BaseRapport(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.dossier = mod_dossier.construire(projet_complet())
        cls.repertoire = tempfile.mkdtemp(prefix="hydrobassin_")

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.repertoire, ignore_errors=True)

    def chemin(self, nom: str) -> str:
        return os.path.join(self.repertoire, nom)


class TestDossier(BaseRapport):
    def test_contenu_du_dossier(self):
        d = self.dossier
        self.assertEqual(len(d.resultats), 4)
        self.assertIsNotNone(d.simulation)
        self.assertIsNotNone(d.table)
        self.assertIsNotNone(d.orifice)
        self.assertGreater(d.resultat_principal.volume_m3, 0)

    def test_synthese(self):
        lignes = mod_dossier.synthese_scenarios(self.dossier)
        self.assertEqual(len(lignes), 5)
        self.assertEqual(len(lignes[0]), 8)

    def test_graphiques(self):
        for g in (self.dossier.graphique_dimensionnement(), self.dossier.graphique_simulation(),
                  self.dossier.graphique_debits(), self.dossier.graphique_orifice()):
            self.assertIsNotNone(g)
            self.assertTrue(g.series)
            png = charts.rendre_png(g, 400, 220)
            self.assertEqual(png[:8], b"\x89PNG\r\n\x1a\n")
            largeur, hauteur = struct.unpack(">II", png[16:24])
            self.assertEqual((largeur, hauteur), (400, 220))

    def test_dossier_sans_bassin(self):
        projet = projet_complet()
        projet.bassin = Bassin()
        d = mod_dossier.construire(projet)
        self.assertIsNone(d.simulation)
        self.assertIsNone(d.table)


class TestExcel(BaseRapport):
    def test_generation(self):
        import openpyxl

        chemin = xlsx_report.ecrire(self.dossier, self.chemin("rapport.xlsx"))
        self.assertTrue(os.path.getsize(chemin) > 10000)
        wb = openpyxl.load_workbook(chemin)
        self.assertEqual(
            wb.sheetnames,
            ["Projet", "Pluie de projet", "Scénarios", "Bassin - table QDF", "Ajutage",
             "Pluies statistiques"],
        )
        noms = set(wb.defined_names)
        for attendu in ("S_ponderee", "Q_infiltration", "Q_ajutage", "V_bassin", "V_sous_ajutage"):
            self.assertIn(attendu, noms)

    def test_le_classeur_contient_des_formules(self):
        import openpyxl

        chemin = xlsx_report.ecrire(self.dossier, self.chemin("formules.xlsx"))
        wb = openpyxl.load_workbook(chemin)
        ws = wb["Pluie de projet"]
        formules = [c.value for ligne in ws.iter_rows(min_row=9, max_row=12) for c in ligne
                    if isinstance(c.value, str) and c.value.startswith("=")]
        self.assertGreater(len(formules), 10)
        self.assertTrue(any("S_ponderee" in f for f in formules))
        ws = wb["Scénarios"]
        self.assertTrue(str(ws["B5"].value).startswith("=MAX("))

    @unittest.skipUnless(os.environ.get("HYDROBASSIN_TEST_FORMULES"),
                         "évaluation des formules Excel (variable HYDROBASSIN_TEST_FORMULES)")
    def test_les_formules_reproduisent_le_moteur(self):
        """Recalcule le classeur avec la bibliothèque `formulas` et compare au moteur Python."""
        import formulas  # dépendance de test uniquement

        chemin = xlsx_report.ecrire(self.dossier, self.chemin("verif.xlsx"))
        modele = formulas.ExcelModel().loads(chemin).finish()
        solution = modele.calculate()
        base = os.path.basename(chemin).upper()

        def valeur(feuille, cellule):
            cle = f"'[{base}]{feuille}'!{cellule}"
            for k, v in solution.items():
                if k.upper().endswith(cle):
                    try:
                        return float(v.value[0, 0])
                    except Exception:
                        return v
            raise KeyError(cle)

        for colonne, scenario in zip("BCDE", mod_dossier.ORDRE_SCENARIOS):
            attendu = self.dossier.resultats[scenario]
            self.assertAlmostEqual(valeur("SCÉNARIOS", f"{colonne}5"), attendu.volume_m3, places=6)
            self.assertAlmostEqual(valeur("SCÉNARIOS", f"{colonne}6"), attendu.duree_critique_min, places=6)
            self.assertAlmostEqual(valeur("SCÉNARIOS", f"{colonne}12"), attendu.temps_vidange_h, places=6)
        self.assertAlmostEqual(valeur("AJUTAGE", "B10"), self.dossier.orifice.diametre_mm, places=6)


class TestExcelQDF(unittest.TestCase):
    """Avec les tables QDF, le classeur doit balayer les mêmes durées que l'application."""

    @classmethod
    def setUpClass(cls):
        p = projet_complet()
        p.source_pluie = "qdf"
        cls.dossier = mod_dossier.construire(p)
        cls.repertoire = tempfile.mkdtemp(prefix="hydrobassin_qdf_")

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.repertoire, ignore_errors=True)

    def test_le_classeur_ne_balaie_que_les_durees_tabulees(self):
        import openpyxl

        chemin = xlsx_report.ecrire(self.dossier, os.path.join(self.repertoire, "qdf.xlsx"))
        classeur = openpyxl.load_workbook(chemin)
        feuille = classeur["Pluie de projet"]
        durees = [c.value for c in feuille["A"][8:] if isinstance(c.value, (int, float))]
        self.assertEqual(sorted(durees), sorted(float(d) for d in rainfall.QDF_DURATIONS_MIN))

    @unittest.skipUnless(os.environ.get("HYDROBASSIN_TEST_FORMULES"),
                         "évaluation des formules Excel (variable HYDROBASSIN_TEST_FORMULES)")
    def test_les_formules_qdf_reproduisent_le_moteur(self):
        import formulas

        chemin = xlsx_report.ecrire(self.dossier, os.path.join(self.repertoire, "verif_qdf.xlsx"))
        modele = formulas.ExcelModel().loads(chemin).finish()
        solution = modele.calculate()
        base = os.path.basename(chemin).upper()

        def valeur(feuille, cellule):
            cle = f"'[{base}]{feuille}'!{cellule}"
            for k, v in solution.items():
                if k.upper().endswith(cle):
                    try:
                        return float(v.value[0, 0])
                    except Exception:
                        return v
            raise KeyError(cle)

        for colonne, scenario in zip("BCDE", mod_dossier.ORDRE_SCENARIOS):
            attendu = self.dossier.resultats[scenario]
            self.assertAlmostEqual(valeur("SCÉNARIOS", f"{colonne}5"), attendu.volume_m3, places=6)
            self.assertAlmostEqual(valeur("SCÉNARIOS", f"{colonne}6"),
                                   attendu.duree_critique_min, places=6)


class TestWord(BaseRapport):
    def test_generation(self):
        chemin = docx_report.ecrire(self.dossier, self.chemin("rapport.docx"))
        with zipfile.ZipFile(chemin) as z:
            noms = z.namelist()
            self.assertIn("word/document.xml", noms)
            self.assertIn("[Content_Types].xml", noms)
            self.assertGreaterEqual(len([n for n in noms if n.startswith("word/media/")]), 3)
            document = z.read("word/document.xml").decode("utf-8")
        ET.fromstring(document)  # XML bien formé
        self.assertIn("bassin d'orage", document)
        self.assertIn("Bütgenbach", document)

    def test_les_images_sont_referencees(self):
        chemin = docx_report.ecrire(self.dossier, self.chemin("images.docx"))
        with zipfile.ZipFile(chemin) as z:
            rels = z.read("word/_rels/document.xml.rels").decode()
            document = z.read("word/document.xml").decode()
            medias = [n.split("/")[-1] for n in z.namelist() if n.startswith("word/media/")]
        for media in medias:
            self.assertIn(media, rels)
        for i in range(1, len(medias) + 1):
            self.assertIn(f'r:embed="rIdImg{i}"', document)


class TestRapportAvecAmont(unittest.TestCase):
    """Les trois formats documentent le bassin d'orage amont."""

    @classmethod
    def setUpClass(cls):
        p = projet_complet()
        p.amont = BassinAmont(actif=True, surface_bv_m2=8000.0, coef_ruissellement=0.8,
                              debit_ajutage_ls=2.0, surface_dispersion_m2=100.0,
                              k_infiltration_ms=1e-5, volume_temporisation_m3=300.0,
                              inclure_bv_dans_ajutage=True)
        cls.dossier = mod_dossier.construire(p)
        cls.repertoire = tempfile.mkdtemp(prefix="hydrobassin_amont_")

    @classmethod
    def tearDownClass(cls):
        shutil.rmtree(cls.repertoire, ignore_errors=True)

    def chemin(self, nom: str) -> str:
        return os.path.join(self.repertoire, nom)

    def test_le_bassin_amont_figure_meme_sans_ouvrage_encode(self):
        """C'est une donnée d'entrée : elle ne doit pas dépendre de l'ouvrage aval."""
        projet = projet_complet()
        projet.bassin = Bassin()          # aucun ouvrage encodé
        projet.amont = BassinAmont(actif=True, surface_bv_m2=10000.0,
                                   coef_ruissellement=0.8, debit_ajutage_ls=2.0,
                                   volume_temporisation_m3=120.0)
        sans_ouvrage = mod_dossier.construire(projet)
        self.assertIsNone(sans_ouvrage.simulation)
        chemin = docx_report.ecrire(sans_ouvrage, self.chemin("amont_sans_ouvrage.docx"))
        with zipfile.ZipFile(chemin) as z:
            document = z.read("word/document.xml").decode("utf-8")
        self.assertIn("1.3 Bassin d", document)
        self.assertIn("bassin versant amont", document)
        self.assertIn("sous-dimensionn", document)

    def test_word_decrit_le_bassin_amont(self):
        chemin = docx_report.ecrire(self.dossier, self.chemin("amont.docx"))
        with zipfile.ZipFile(chemin) as z:
            document = z.read("word/document.xml").decode("utf-8")
        ET.fromstring(document)
        self.assertIn("bassin versant amont", document)
        self.assertIn("Volume de temporisation amont", document)

    def test_pdf_decrit_le_bassin_amont(self):
        chemin = pdf_report.ecrire(self.dossier, self.chemin("amont.pdf"))
        self.assertGreater(os.path.getsize(chemin), 5000)
        with open(chemin, "rb") as fh:
            self.assertTrue(fh.read().startswith(b"%PDF"))

    def test_excel_compte_la_surface_amont(self):
        import openpyxl

        chemin = xlsx_report.ecrire(self.dossier, self.chemin("amont.xlsx"))
        classeur = openpyxl.load_workbook(chemin)
        libelles = [c.value for ligne in classeur["Projet"].iter_rows()
                    for c in ligne if isinstance(c.value, str)]
        self.assertTrue(any("Bassin d'orage amont" in (l or "") for l in libelles))
        self.assertTrue(any("BV amont compris" in (l or "") for l in libelles))


    def test_excel_annonce_le_volume_avec_l_apport_amont(self):
        """Les formules vives ignorent l'amont : le classeur doit le dire et donner
        la valeur qui fait foi, sinon il contredit le rapport PDF."""
        import openpyxl

        chemin = xlsx_report.ecrire(self.dossier, self.chemin("amont_volume.xlsx"))
        feuille = openpyxl.load_workbook(chemin)["Scénarios"]
        textes = [c.value for ligne in feuille.iter_rows()
                  for c in ligne if isinstance(c.value, str)]
        self.assertTrue(any("intégration pas à pas" in (t or "") for t in textes))
        ligne = next(l for l in feuille.iter_rows()
                     if isinstance(l[0].value, str)
                     and l[0].value.startswith("Volume à maîtriser, apport du bassin amont"))
        valeurs = [c.value for c in ligne[1:5]]
        attendus = [round(self.dossier.resultats[s].volume_m3, 1) for s in mod_dossier.ORDRE_SCENARIOS]
        self.assertEqual(valeurs, attendus)


class TestVirguleDecimale(BaseRapport):

    """Les rapports francophones affichent une virgule décimale, pas un point."""

    #: Les numéros de section (« 1.1 Surfaces incidentes ») gardent leur point.
    TITRE = re.compile(r"^\d+(?:\.\d+)* ")

    def _fautifs(self, textes):
        return [t for t in textes
                if re.search(r"\d\.\d", t) and not self.TITRE.match(t.strip())]

    def test_word(self):
        chemin = docx_report.ecrire(self.dossier, self.chemin("virgules.docx"))
        with zipfile.ZipFile(chemin) as z:
            document = z.read("word/document.xml").decode("utf-8")
        textes = re.findall(r"<w:t[^>]*>(.*?)</w:t>", document, re.S)
        self.assertTrue(any(re.search(r"\d,\d", t) for t in textes))
        self.assertEqual(self._fautifs(textes), [])

    def test_pdf(self):
        chemin = pdf_report.ecrire(self.dossier, self.chemin("virgules.pdf"))
        with open(chemin, "rb") as fh:
            brut = fh.read()
        flux = []
        for bloc in re.finditer(rb"stream\r?\n(.*?)\r?\nendstream", brut, re.S):
            try:
                flux.append(zlib.decompress(bloc.group(1)).decode("latin-1"))
            except zlib.error:
                continue
        textes = [t[1:-1] for t in re.findall(r"\((?:[^()\\]|\\.)*\)", "\n".join(flux))]
        self.assertTrue(any(re.search(r"\d,\d", t) for t in textes))
        self.assertEqual(self._fautifs(textes), [])

    def test_graduations_de_graphique(self):
        self.assertEqual(charts.format_nombre(66.3), "66,3")
        self.assertEqual(charts.format_nombre(1200), "1200")
        self.assertIn(",", charts._FONT)  # la police 5x7 sait tracer la virgule


class TestPdf(BaseRapport):
    def test_generation(self):
        chemin = pdf_report.ecrire(self.dossier, self.chemin("rapport.pdf"))
        with open(chemin, "rb") as fh:
            data = fh.read()
        self.assertTrue(data.startswith(b"%PDF-1.4"))
        self.assertTrue(data.rstrip().endswith(b"%%EOF"))
        self.assertGreater(data.count(b"/Type /Page"), 3)
        self.assertIn(b"startxref", data)

    def test_moteur_pdf(self):
        pdf = Pdf()
        pdf.pied = "test"
        pdf.titre("Titre é à ç")
        pdf.texte("Un paragraphe " * 40)
        pdf.tableau([["a", "b"], ["1", "2"]], [200, 200])
        pdf.encadre("Encadré")
        pdf.nouvelle_page()
        pdf.texte("page 2")
        chemin = pdf.enregistrer(self.chemin("moteur.pdf"))
        with open(chemin, "rb") as fh:
            data = fh.read()
        self.assertEqual(data.count(b"/Type /Page "), 2)

    def test_nettoyage_hors_winansi(self):
        self.assertEqual(nettoyer("cœur ≤ 3"), "coeur <= 3")
        self.assertNotIn("œ", nettoyer("œuf"))

    def test_largeur_texte(self):
        self.assertGreater(largeur_texte("MMMM", 10), largeur_texte("iiii", 10))
        self.assertAlmostEqual(largeur_texte("é", 10), largeur_texte("e", 10))


class TestGraphiques(unittest.TestCase):
    def test_graduations(self):
        self.assertEqual(charts.graduations(0, 10, 5), [0, 2, 4, 6, 8, 10])
        self.assertTrue(charts.graduations(0, 1, 4))

    def test_format(self):
        self.assertEqual(charts.format_duree_courte(30), "30min")
        self.assertEqual(charts.format_duree_courte(120), "2h")
        self.assertEqual(charts.format_duree_courte(2880), "2j")
        self.assertEqual(charts.format_nombre(0), "0")

    def test_png_valide(self):
        g = charts.Graphique(
            axe_x="Durée de pluie", axe_y="V",
            series=[charts.Serie("v", [(10, 1), (100, 5), (1000, 2)], charts.BLEU, aire=True)],
            reperes=[charts.Repere(3, "seuil"), charts.Repere(100, "t", vertical=True)],
            x_log=True,
        )
        png = charts.rendre_png(g, 300, 180)
        self.assertEqual(png[:8], b"\x89PNG\r\n\x1a\n")
        self.assertIn("v", charts.legende_texte(g))

    def test_graphique_vide(self):
        png = charts.rendre_png(charts.Graphique(), 120, 80)
        self.assertEqual(png[:8], b"\x89PNG\r\n\x1a\n")


if __name__ == "__main__":
    unittest.main(verbosity=2)
