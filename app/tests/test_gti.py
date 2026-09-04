"""Conformité à la fiche de calcul officielle du GTI (Service public de Wallonie).

Deux niveaux de vérification :

* un jeu de cas de référence figé (``donnees_gti.json``), calculé par la chaîne
  de la fiche GTI et rejoué à chaque exécution ;
* si le classeur GTI est disponible, une campagne complète : tous les
  coefficients de Montana, puis un balayage communes x récurrences x scénarios.

La fiche officielle procède ainsi (feuille « Pluie », colonnes A a K) :

    a, b   = IF(t<25, a1, IF(t<=6000, a2, a3))
    i      = a * t^(-b)                      [mm/h]
    h      = i * t / 60                      [mm]
    V_in   = h * S_ponderee / 1000           [m³]
    V_out  = Q * t * 60 / 1000               [m³]
    V      = MAX(V_in - V_out, 0)            [m³]

le volume retenu étant le maximum sur les 17 280 durées de 10 a 86 405 minutes.
En « infiltration seule », Q est le seul débit d'infiltration ; en
« infiltration et rejet », Q est la somme de l'infiltration et du rejet
admissible (5 l/s/ha de surface raccordée).
"""

import json
import os
import sys
import unittest

sys.path.insert(0, os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "src"))

from bassin.core import hydro, rainfall  # noqa: E402
from bassin.core.model import (  # noqa: E402
    COEF_SECURITE_INFILTRATION,
    Projet,
    SCENARIO_DISPERSION,
    SCENARIO_MIXTE,
    debit_infiltration_ls,
)

DONNEES = os.path.join(os.path.dirname(os.path.abspath(__file__)), "donnees_gti.json")
#: Classeur fourni par l'utilisateur ; absent de l'intégration continue.
CLASSEUR = os.environ.get(
    "HYDROBASSIN_CLASSEUR_GTI",
    "/root/.claude/uploads/56ed6976-800a-50d9-bd1d-a50d9f69108f/657e4413-GTI_deverrouille.xlsx",
)


def projet_pour(ins: str, nom: str, recurrence: int, surface_ponderee_m2: float,
                surface_totale_m2: float = 0.0) -> Projet:
    """Projet dont les surfaces pondérée et totale valent exactement les valeurs voulues.

    La fiche GTI distingue les deux : la surface pondérée alimente le
    ruissellement, la surface totale fixe le rejet admissible de 5 l/s/ha. On
    répartit donc entre toitures (coefficient 1,0) et prairies (0,15).
    """
    p = Projet(commune_ins=ins, commune_nom=nom, periode_retour=recurrence,
               surfaces=Projet.surfaces_par_defaut(), source_pluie="montana")
    if surface_totale_m2 <= surface_ponderee_m2:
        p.surfaces[7].aire_m2 = surface_ponderee_m2      # coefficient 1,0
        return p
    coef_prairie = p.surfaces[1].coefficient              # 0,15
    prairies = (surface_totale_m2 - surface_ponderee_m2) / (1.0 - coef_prairie)
    p.surfaces[1].aire_m2 = prairies
    p.surfaces[7].aire_m2 = surface_totale_m2 - prairies
    return p


class TestConformiteGTI(unittest.TestCase):
    """Cas de référence figés, rejoués sans dépendre du classeur."""

    @classmethod
    def setUpClass(cls):
        with open(DONNEES, encoding="utf-8") as fh:
            cls.reference = json.load(fh)

    def test_la_grille_de_durees_est_celle_de_la_fiche(self):
        """La feuille « Pluie » compte 17 280 lignes, de 10 a 86 405 min."""
        durees = rainfall.SourcePluie("63013", 25, rainfall.SOURCE_MONTANA).durees_de_balayage(
            hydro.DUREE_MIN, hydro.DUREE_MAX, hydro.PAS_DUREE)
        self.assertEqual(len(durees), 17280)
        self.assertEqual(durees[0], 10.0)
        self.assertEqual(durees[-1], 86405.0)

    def test_les_cas_de_reference_sont_reproduits(self):
        for cas in self.reference["cas"]:
            with self.subTest(commune=cas["commune"], recurrence=cas["recurrence"],
                              scenario=cas["scenario"]):
                p = projet_pour(cas["ins"], cas["commune"], cas["recurrence"],
                                cas["surface_ponderee_m2"])
                serie = hydro.serie_projet(p)
                volume, duree, _hauteur = hydro.volume_a_maitriser(
                    serie, p.aire_ponderee_m2, cas["debit_sortie_ls"])
                self.assertAlmostEqual(volume, cas["volume_m3"], places=4)
                self.assertEqual(duree, cas["duree_min"])

    def test_le_debit_d_infiltration_suit_la_formule_de_la_fiche(self):
        """« Infiltration seule »!B36 = 1000 * S * K / 2 : le facteur 2 est celui du GTI."""
        self.assertEqual(COEF_SECURITE_INFILTRATION, 2.0)
        for cas in self.reference["cas"]:
            with self.subTest(commune=cas["commune"]):
                attendu = 1000.0 * cas["surface_infiltrante_m2"] * cas["k_infiltration_ms"] / 2.0
                self.assertAlmostEqual(attendu, cas["debit_infiltration_ls"], places=9)
                self.assertAlmostEqual(
                    debit_infiltration_ls(cas["surface_infiltrante_m2"],
                                          cas["k_infiltration_ms"]),
                    cas["debit_infiltration_ls"], places=9)

    def test_le_rejet_admissible_suit_la_formule_de_la_fiche(self):
        """« Infiltration et rejet »!B38 = 5 l/s/ha de surface raccordée (non pondérée)."""
        for cas in self.reference["cas"]:
            if cas["scenario"] != "infiltration et rejet":
                continue
            with self.subTest(commune=cas["commune"]):
                p = projet_pour(cas["ins"], cas["commune"], cas["recurrence"],
                                cas["surface_ponderee_m2"], cas["surface_totale_m2"])
                self.assertAlmostEqual(p.debit_fuite_admissible_ls, cas["debit_rejet_ls"],
                                       places=9)

    def test_les_deux_scenarios_officiels_passent_par_le_moteur_complet(self):
        """Les entrées de la fiche (surface infiltrante, K) traversent tout le moteur."""
        for cas in self.reference["cas"]:
            with self.subTest(commune=cas["commune"], recurrence=cas["recurrence"],
                              scenario=cas["scenario"]):
                p = projet_pour(cas["ins"], cas["commune"], cas["recurrence"],
                                cas["surface_ponderee_m2"], cas["surface_totale_m2"])
                self.assertAlmostEqual(p.aire_ponderee_m2, cas["surface_ponderee_m2"], places=6)
                self.assertAlmostEqual(p.aire_totale_m2, cas["surface_totale_m2"], places=6)
                p.surface_infiltration_m2 = cas["surface_infiltrante_m2"]
                p.k_infiltration_ms = cas["k_infiltration_ms"]
                if cas["scenario"] == "infiltration seule":
                    res = hydro.dimensionner(p, SCENARIO_DISPERSION, avec_minima=False)
                else:
                    # La fiche impose le rejet admissible de 5 l/s/ha.
                    p.debit_ajutage_ls = p.debit_fuite_admissible_ls
                    res = hydro.dimensionner(p, SCENARIO_MIXTE, avec_minima=False)
                self.assertAlmostEqual(res.debit_infiltration_ls,
                                       cas["debit_infiltration_ls"], places=9)
                self.assertAlmostEqual(res.debit_sortant_ls, cas["debit_sortie_ls"], places=9)
                self.assertAlmostEqual(res.volume_m3, cas["volume_m3"], places=4)
                self.assertEqual(res.duree_critique_min, cas["duree_min"])
                self.assertAlmostEqual(res.intensite_ls_ha, cas["intensite_ls_ha"], places=4)
                self.assertAlmostEqual(res.debit_entrant_ls, cas["debit_entrant_ls"], places=4)
                # Calcul!B40 : la fiche divise le volume arrondi au dm³.
                vidange = round(res.volume_m3, 1) / res.debit_sortant_ls * 1000.0 / 3600.0
                self.assertAlmostEqual(vidange, cas["vidange_h"], places=6)

    def test_les_parametres_de_reference_du_gti_sont_ceux_de_la_fiche(self):
        """Calcul!H3 = 48 h de vidange, H4 = 10 % de la surface de référence."""
        from bassin.core.model import (
            DEBIT_FUITE_SPECIFIQUE_MAX_LS_HA,
            POURCENTAGE_SURFACE_INFILTRATION_LIMITE,
            TEMPS_VIDANGE_LIMITE_H,
        )

        self.assertEqual(TEMPS_VIDANGE_LIMITE_H, 48)
        self.assertEqual(POURCENTAGE_SURFACE_INFILTRATION_LIMITE, 0.1)
        self.assertEqual(DEBIT_FUITE_SPECIFIQUE_MAX_LS_HA, 5)


@unittest.skipUnless(os.path.exists(CLASSEUR), f"classeur GTI absent ({CLASSEUR})")
class TestClasseurGTI(unittest.TestCase):
    """Campagne complète contre le classeur officiel, quand il est disponible."""

    @classmethod
    def setUpClass(cls):
        import openpyxl

        classeur = openpyxl.load_workbook(CLASSEUR, data_only=True, read_only=True)
        cls.coefficients = {}
        for ligne in classeur["Montana"].iter_rows(min_row=2, values_only=True):
            if ligne[1] is None:
                continue
            _, ins, _nom, rp = ligne[:4]
            cls.coefficients[(f"{int(ins):05d}", int(rp))] = tuple(ligne[4:10])
        classeur.close()

    def test_tous_les_coefficients_de_montana_sont_identiques(self):
        """Le référentiel embarqué doit être celui du classeur, sans exception."""
        for (ins, rp), attendu in self.coefficients.items():
            obtenu = rainfall.montana_coeffs(ins, rp)
            self.assertEqual(len(obtenu), 6)
            for a, b in zip(attendu, obtenu):
                self.assertAlmostEqual(a, b, places=9, msg=f"INS={ins} T={rp}")

    def test_aucune_commune_inventee(self):
        for commune in rainfall.communes():
            if not commune.a_montana:
                continue
            for rp in rainfall.RETURN_PERIODS:
                self.assertIn((commune.ins, rp), self.coefficients)

    def test_le_dimensionnement_suit_la_fiche(self):
        """Balayage communes x récurrences x scénarios contre la chaîne officielle."""
        def volume_gti(ins, rp, s_pond, q_out):
            a1, b1, a2, b2, a3, b3 = self.coefficients[(ins, rp)]
            meilleur = (0.0, 10.0)
            for t in range(10, 86406, 5):
                if t < 25:
                    a, b = a1, b1
                elif t <= 6000:
                    a, b = a2, b2
                else:
                    a, b = a3, b3
                h = a * t ** (-b) * t / 60.0
                v = max(h * s_pond / 1000.0 - q_out * t * 60.0 / 1000.0, 0.0)
                if v > meilleur[0]:
                    meilleur = (v, float(t))
            return meilleur

        communes = [("63013", "Bütgenbach"), ("81001", "Arlon"), ("52011", "Charleroi"),
                    ("62063", "Liège"), ("92094", "Namur")]
        for ins, nom in communes:
            for rp in (10, 25, 100, 200):
                for s_pond, s_inf, k in ((1575.0, 120.0, 1e-5), (10000.0, 400.0, 5e-6)):
                    p = projet_pour(ins, nom, rp, s_pond)
                    p.surface_infiltration_m2, p.k_infiltration_ms = s_inf, k
                    q_inf = debit_infiltration_ls(s_inf, k, p.coef_securite_infiltration)
                    with self.subTest(commune=nom, recurrence=rp, scenario="infiltration seule"):
                        res = hydro.dimensionner(p, SCENARIO_DISPERSION, avec_minima=False)
                        attendu, duree = volume_gti(ins, rp, p.aire_ponderee_m2, q_inf)
                        self.assertAlmostEqual(res.volume_m3, attendu, places=6)
                        self.assertEqual(res.duree_critique_min, duree)
                    with self.subTest(commune=nom, recurrence=rp, scenario="infiltration et rejet"):
                        # La fiche impose le rejet admissible de 5 l/s/ha.
                        p.debit_ajutage_ls = 5.0 * p.aire_totale_m2 / 10000.0
                        res = hydro.dimensionner(p, SCENARIO_MIXTE, avec_minima=False)
                        attendu, duree = volume_gti(ins, rp, p.aire_ponderee_m2,
                                                    q_inf + p.debit_ajutage_ls)
                        self.assertAlmostEqual(res.volume_m3, attendu, places=6)
                        self.assertEqual(res.duree_critique_min, duree)

    def test_les_intensites_correspondent_aux_valeurs_en_cache(self):
        """Le classeur conserve les intensités calculées par Excel : on s'y compare."""
        import openpyxl

        classeur = openpyxl.load_workbook(CLASSEUR, data_only=True, read_only=True)
        pluie = classeur["Pluie"]
        lignes = list(pluie.iter_rows(min_row=11, max_row=17290, min_col=3, max_col=5,
                                      values_only=True))
        classeur.close()
        compares = 0
        for duree, intensite, hauteur in lignes:
            if duree is None or intensite is None:
                continue
            self.assertAlmostEqual(rainfall.intensite_montana("63013", 25, duree), intensite,
                                   places=9, msg=f"durée {duree} min")
            if hauteur is not None:
                self.assertAlmostEqual(rainfall.hauteur_montana("63013", 25, duree), hauteur,
                                       places=9, msg=f"durée {duree} min")
            compares += 1
        self.assertGreater(compares, 17000)


if __name__ == "__main__":
    unittest.main()
