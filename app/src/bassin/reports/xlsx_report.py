"""Rapport Excel (.xlsx) avec formules de calcul vivantes.

Le classeur génère n'est pas une simple exportation de résultats : les feuilles
"Pluie de projet", "Scénarios", "Bassin" et "Ajutage" contiennent les formules
de la méthode rationnelle. L'utilisateur peut modifier les données d'entrée
(surfaces, K, débit d'ajutage, volumes) et voir les résultats se recalculer.
"""

from __future__ import annotations

import math
from typing import Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

from ..core import rainfall
from ..core.model import (
    LIBELLES_SCENARIOS,
    SCENARIO_DISPERSION,
    SCENARIO_MIXTE,
    SCENARIO_SEUIL,
    SCENARIO_TEMPORISATION,
)
from .dossier import Dossier, ORDRE_SCENARIOS

BLEU = "1D4ED8"
BLEU_PALE = "DBEAFE"
GRIS_PALE = "F1F5F9"
VERT_PALE = "DCFCE7"
ROUGE_PALE = "FEE2E2"
ORANGE_PALE = "FEF3C7"
BLANC = "FFFFFF"

_BORDURE = Border(*[Side(style="thin", color="CBD5E1")] * 4)


def _titre(ws, cellule: str, texte: str, taille: int = 14) -> None:
    ws[cellule] = texte
    ws[cellule].font = Font(bold=True, size=taille, color=BLEU)


def _entete(ws, ligne: int, valeurs: Sequence[str], col_debut: int = 1) -> None:
    for i, v in enumerate(valeurs):
        c = ws.cell(row=ligne, column=col_debut + i, value=v)
        c.font = Font(bold=True, color=BLANC)
        c.fill = PatternFill("solid", fgColor=BLEU)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _BORDURE


def _label(ws, ligne: int, texte: str, valeur=None, unite: str = "", format_nombre: Optional[str] = None,
           col: int = 1, gras: bool = False, fond: Optional[str] = None):
    c0 = ws.cell(row=ligne, column=col, value=texte)
    c0.font = Font(bold=gras)
    c1 = ws.cell(row=ligne, column=col + 1, value=valeur)
    if format_nombre:
        c1.number_format = format_nombre
    c1.font = Font(bold=True)
    if fond:
        c1.fill = PatternFill("solid", fgColor=fond)
    ws.cell(row=ligne, column=col + 2, value=unite)
    return c1


def _largeurs(ws, largeurs: Dict[str, int]) -> None:
    for col, w in largeurs.items():
        ws.column_dimensions[col].width = w


def _grille_durees(dossier: Dossier) -> List[float]:
    """Durées balayées par le classeur, calquées sur celles de l'application.

    Avec les tables QDF, seules les 19 durées normalisées ont un sens : ajouter
    une grille logarithmique ferait retenir au classeur une durée critique
    interpolée, différente de celle affichée par l'application.
    """
    durees = set(float(d) for d in rainfall.QDF_DURATIONS_MIN)
    src = rainfall.SourcePluie(dossier.projet.commune_ins, dossier.projet.periode_retour,
                               dossier.projet.source_pluie)
    if not src.durees_tabulees:
        for i in range(101):
            durees.add(round(10 * (8640 ** (i / 100.0)), 1))
        for r in dossier.resultats.values():
            if r.duree_critique_min:
                durees.add(float(r.duree_critique_min))
        if dossier.duree_critique_min:
            durees.add(float(dossier.duree_critique_min))
    return sorted(d for d in durees if 10 <= d <= 86400)


def construire_classeur(dossier: Dossier) -> Workbook:
    projet = dossier.projet
    wb = Workbook()

    # ------------------------------------------------------------------ Projet
    ws = wb.active
    ws.title = "Projet"
    _largeurs(ws, {"A": 46, "B": 16, "C": 16, "D": 18, "E": 30})
    _titre(ws, "A1", "DIMENSIONNEMENT D'UN BASSIN D'ORAGE", 16)
    ws["A2"] = "Méthode rationnelle - pluies statistiques du GTI (Région wallonne)"
    ws["A2"].font = Font(italic=True, color="475569")

    _label(ws, 4, "Projet", projet.nom_projet or "-", gras=True)
    _label(ws, 5, "Localisation", projet.localisation or "-")
    _label(ws, 6, "Auteur", projet.auteur or "-")
    _label(ws, 7, "Date", dossier.date)
    _label(ws, 8, "Commune", projet.commune_nom)
    _label(ws, 9, "Code INS", projet.commune_ins)
    _label(ws, 10, "Période de retour", projet.periode_retour, "ans")
    _label(ws, 11, "Source des pluies", dossier.libelle_source)

    _titre(ws, "A13", "1. Surfaces incidentes", 12)
    _entete(ws, 14, ["Type d'occupation du sol", "Coeff. ruiss. [-]", "Surface [m²]",
                     "Surface pondérée [m²]", "Notes"])
    ligne = 15
    for s in projet.surfaces:
        ws.cell(row=ligne, column=1, value=s.libelle).border = _BORDURE
        ws.cell(row=ligne, column=2, value=s.coefficient).border = _BORDURE
        ws.cell(row=ligne, column=3, value=s.aire_m2).border = _BORDURE
        c = ws.cell(row=ligne, column=4, value=f"=B{ligne}*C{ligne}")
        c.border = _BORDURE
        c.number_format = "0.0"
        ws.cell(row=ligne, column=5, value=s.note).border = _BORDURE
        ligne += 1
    l_tot = ligne
    ws.cell(row=l_tot, column=1, value="TOTAL").font = Font(bold=True)
    for col, formule in ((3, f"=SUM(C15:C{l_tot - 1})"), (4, f"=SUM(D15:D{l_tot - 1})")):
        c = ws.cell(row=l_tot, column=col, value=formule)
        c.font = Font(bold=True)
        c.number_format = "0.0"
        c.fill = PatternFill("solid", fgColor=BLEU_PALE)
        c.border = _BORDURE
    l_coef = l_tot + 1
    c = ws.cell(row=l_coef, column=1, value="Coefficient de ruissellement moyen")
    c.font = Font(bold=True)
    c = ws.cell(row=l_coef, column=4, value=f"=IF(C{l_tot}>0,D{l_tot}/C{l_tot},0)")
    c.number_format = "0.000"
    c.font = Font(bold=True)
    c.fill = PatternFill("solid", fgColor=BLEU_PALE)

    l = l_coef + 2
    _titre(ws, f"A{l}", "2. Sol, exutoire et contraintes", 12)
    l += 1
    c_sref = _label(ws, l, "Surface de référence du projet", projet.surface_reference_m2, "m²", "0.0"); l += 1
    c_k = _label(ws, l, "Coefficient d'infiltration K", projet.k_infiltration_ms, "m/s", "0.00E+00"); l += 1
    c_cs = _label(ws, l, "Coefficient de sécurité sur K", projet.coef_securite_infiltration, "[-]", "0.0"); l += 1
    c_sinf = _label(ws, l, "Surface d'infiltration du dispositif", projet.surface_infiltration_m2, "m²", "0.0"); l += 1
    c_qaj = _label(ws, l, "Débit d'ajutage (orifice calibré)", projet.debit_ajutage_ls, "l/s", "0.000"); l += 1
    c_tvid = _label(ws, l, "Temps de vidange maximum admis (après la pluie)", projet.temps_vidange_max_h, "h", "0.0"); l += 1
    c_qinf = _label(ws, l, "Débit d'infiltration Q = 1000.S.K/coef",
                    f"=1000*{c_sinf.coordinate}*{c_k.coordinate}/{c_cs.coordinate}", "l/s", "0.000",
                    fond=VERT_PALE)
    l_qinf = l; l += 1
    # Un bassin versant amont raccordé compte dans la surface qui fixe le débit
    # de fuite admissible, si l'utilisateur a demandé de le prendre en compte.
    amont = projet.amont
    if amont.actif and amont.inclure_bv_dans_ajutage:
        c_qadm = _label(ws, l, "Débit de fuite admissible (5 l/s/ha, BV amont compris)",
                        f"=5*(C{l_tot}+{amont.surface_bv_m2})/10000", "l/s", "0.000"); l += 2
    else:
        c_qadm = _label(ws, l, "Débit de fuite admissible (5 l/s/ha)",
                        f"=5*C{l_tot}/10000", "l/s", "0.000"); l += 2

    if amont.actif:
        _titre(ws, f"A{l}", "2 bis. Bassin d'orage amont", 12)
        l += 1
        c_sam = _label(ws, l, "Surface du bassin versant amont", amont.surface_bv_m2, "m²", "0"); l += 1
        c_cam = _label(ws, l, "Coefficient de ruissellement moyen amont", amont.coef_ruissellement,
                       "[-]", "0.00"); l += 1
        _label(ws, l, "Surface active amont",
               f"={c_sam.coordinate}*{c_cam.coordinate}", "m²", "0.0", fond=VERT_PALE); l += 1
        _label(ws, l, "Volume de temporisation amont", amont.volume_temporisation_m3, "m³", "0.0"); l += 1
        c_sdam = _label(ws, l, "Surface de dispersion amont", amont.surface_dispersion_m2,
                        "m²", "0.0"); l += 1
        c_kam = _label(ws, l, "Vitesse d'infiltration amont", amont.k_infiltration_ms,
                       "m/s", "0.00E+00"); l += 1
        c_qajam = _label(ws, l, "Débit d'ajutage amont", amont.debit_ajutage_ls, "l/s", "0.000"); l += 1
        _label(ws, l, "Débit d'infiltration amont",
               f"=1000*{c_sdam.coordinate}*{c_kam.coordinate}/{c_cs.coordinate}", "l/s", "0.000",
               fond=VERT_PALE); l += 1
        _label(ws, l, "Débit restitué vers l'ouvrage aval",
               f"={c_qajam.coordinate}", "l/s", "0.000", fond=VERT_PALE); l += 2

    _titre(ws, f"A{l}", "3. Ouvrage encodé", 12)
    l += 1
    c_vtot = _label(ws, l, "Volume total du bassin", projet.bassin.volume_total_m3, "m³", "0.0"); l += 1
    c_vsous = _label(ws, l, "Volume sous l'axe de l'ajutage", projet.bassin.volume_sous_ajutage_m3, "m³", "0.0"); l += 1
    c_sdisp = _label(ws, l, "Surface de dispersion du bassin", projet.bassin.surface_dispersion_m2, "m²", "0.0"); l += 1
    c_qbas = _label(ws, l, "Débit d'ajutage du bassin", projet.bassin.debit_ajutage_ls, "l/s", "0.000"); l += 1
    c_qinfb = _label(ws, l, "Débit d'infiltration du bassin",
                     f"=1000*{c_sdisp.coordinate}*{c_k.coordinate}/{c_cs.coordinate}", "l/s", "0.000",
                     fond=VERT_PALE); l += 1
    c_h = _label(ws, l, "Charge sur l'ajutage (axe -> trop-plein)", projet.hauteur_charge_m, "m", "0.00"); l += 1
    c_cd = _label(ws, l, "Coefficient de débit Cd", projet.coef_debit_orifice, "[-]", "0.00")

    noms = {
        "S_ponderee": f"Projet!$D${l_tot}",
        "S_totale": f"Projet!$C${l_tot}",
        "K_infiltration": f"Projet!${c_k.column_letter}${c_k.row}",
        "Coef_securite": f"Projet!${c_cs.column_letter}${c_cs.row}",
        "S_infiltration": f"Projet!${c_sinf.column_letter}${c_sinf.row}",
        "Q_infiltration": f"Projet!$B${l_qinf}",
        "Q_ajutage": f"Projet!${c_qaj.column_letter}${c_qaj.row}",
        "T_vidange_max": f"Projet!${c_tvid.column_letter}${c_tvid.row}",
        "V_bassin": f"Projet!${c_vtot.column_letter}${c_vtot.row}",
        "V_sous_ajutage": f"Projet!${c_vsous.column_letter}${c_vsous.row}",
        "S_dispersion": f"Projet!${c_sdisp.column_letter}${c_sdisp.row}",
        "Q_ajutage_bassin": f"Projet!${c_qbas.column_letter}${c_qbas.row}",
        "Q_infiltration_bassin": f"Projet!${c_qinfb.column_letter}${c_qinfb.row}",
        "Charge_orifice": f"Projet!${c_h.column_letter}${c_h.row}",
        "Cd_orifice": f"Projet!${c_cd.column_letter}${c_cd.row}",
    }
    for nom, ref in noms.items():
        wb.defined_names.add(DefinedName(nom, attr_text=ref))

    _feuille_pluie(wb, dossier)
    _feuille_scenarios(wb, dossier)
    _feuille_bassin(wb, dossier)
    _feuille_ajutage(wb, dossier)
    _feuille_statistiques(wb, dossier)
    return wb


def _feuille_pluie(wb: Workbook, dossier: Dossier) -> None:
    projet = dossier.projet
    ws = wb.create_sheet("Pluie de projet")
    _largeurs(ws, {"A": 14, "B": 12, "C": 12, "D": 14, "E": 14, "F": 16, "G": 16, "H": 18,
                   "I": 16, "J": 18, "K": 16, "L": 18, "M": 15, "N": 14, "O": 16, "P": 18})
    _titre(ws, "A1", f"Pluie de projet - {projet.commune_nom} - T = {projet.periode_retour} ans", 14)
    ws["A2"] = dossier.libelle_source
    ws["A2"].font = Font(italic=True, color="475569")

    # Les coefficients de Montana ne servent que si l'utilisateur a choisi cette
    # source : sinon le classeur recalculait des intensités de Montana alors que
    # l'application affichait les mesures QDF.
    src_pluie = rainfall.SourcePluie(projet.commune_ins, projet.periode_retour, projet.source_pluie)
    montana = None
    if src_pluie.source == rainfall.SOURCE_MONTANA:
        montana = rainfall.montana_coeffs(projet.commune_ins, projet.periode_retour)
        _entete(ws, 4, ["Coefficients de Montana", "a1", "b1", "a2", "b2", "a3", "b3"])
        ws.cell(row=5, column=1, value="i [mm/h] = a x t[min]^(-b)")
        for i, v in enumerate(montana):
            ws.cell(row=5, column=2 + i, value=v).number_format = "0.0000"
        ws.cell(row=6, column=1, value="Plages : a1/b1 si t < 25 min | a2/b2 si 25 <= t <= 6000 min | a3/b3 si t > 6000 min")
        ws.cell(row=6, column=1).font = Font(italic=True, size=9, color="475569")

    l0 = 8
    _entete(ws, l0, [
        "Durée [min]", "a", "b", "i [mm/h]", "h [mm]", "V ruisselle [m³]",
        "[1] V évacué [m³]", "[1] V à maîtriser [m³]",
        "[2] V évacué [m³]", "[2] V à maîtriser [m³]",
        "[3] V évacué [m³]", "[3] V à maîtriser [m³]",
        "Q entrant [l/s]", "t seuil [min]", "[4] V évacué [m³]", "[4] V à maîtriser [m³]",
    ])
    ws.freeze_panes = f"A{l0 + 1}"

    durees = _grille_durees(dossier)
    src = rainfall.SourcePluie(projet.commune_ins, projet.periode_retour, projet.source_pluie)
    ligne = l0 + 1
    for d in durees:
        r = ligne
        ws.cell(row=r, column=1, value=d).number_format = "0.0"
        if montana:
            ws.cell(row=r, column=2, value=f"=IF(A{r}<25,$B$5,IF(A{r}<=6000,$D$5,$F$5))").number_format = "0.00"
            ws.cell(row=r, column=3, value=f"=IF(A{r}<25,$C$5,IF(A{r}<=6000,$E$5,$G$5))").number_format = "0.0000"
            ws.cell(row=r, column=4, value=f"=B{r}*A{r}^(-C{r})").number_format = "0.00"
        else:
            ws.cell(row=r, column=4, value=src.intensite_mmh(d)).number_format = "0.00"
        ws.cell(row=r, column=5, value=f"=D{r}*A{r}/60").number_format = "0.00"
        ws.cell(row=r, column=6, value=f"=E{r}*S_ponderee/1000").number_format = "0.00"
        # [1] temporisation seule : ajutage uniquement
        ws.cell(row=r, column=7, value=f"=Q_ajutage*A{r}*60/1000").number_format = "0.00"
        ws.cell(row=r, column=8, value=f"=MAX(F{r}-G{r},0)").number_format = "0.00"
        # [2] dispersion seule : infiltration uniquement
        ws.cell(row=r, column=9, value=f"=Q_infiltration*A{r}*60/1000").number_format = "0.00"
        ws.cell(row=r, column=10, value=f"=MAX(F{r}-I{r},0)").number_format = "0.00"
        # [3] temporisation + dispersion
        ws.cell(row=r, column=11, value=f"=(Q_infiltration+Q_ajutage)*A{r}*60/1000").number_format = "0.00"
        ws.cell(row=r, column=12, value=f"=MAX(F{r}-K{r},0)").number_format = "0.00"
        # [4] dispersion + temporisation au-dela du seuil (ajutage sureleve).
        # M : débit ruisselé entrant, N : instant où le niveau atteint l'axe de l'ajutage.
        ws.cell(row=r, column=13, value=f"=F{r}*1000/(A{r}*60)").number_format = "0.000"
        ws.cell(row=r, column=14,
                value=(f'=IF(V_sous_ajutage<=0,0,IF(M{r}-Q_infiltration<=0,"",'
                       f"V_sous_ajutage*1000/(M{r}-Q_infiltration)/60))")).number_format = "0.0"
        ws.cell(row=r, column=15,
                value=(f'=IF(N{r}="",Q_infiltration*A{r}*60/1000,'
                       f"(Q_infiltration*A{r}+Q_ajutage*MAX(A{r}-N{r},0))*60/1000)")).number_format = "0.00"
        ws.cell(row=r, column=16, value=f"=MAX(F{r}-O{r},0)").number_format = "0.00"
        ligne += 1
    ws["A3"] = f"Plage balayée : {durees[0]:.0f} min a {durees[-1] / 1440:.0f} jours ({len(durees)} durées)"
    ws["A3"].font = Font(italic=True, size=9, color="475569")
    ws._plage_pluie = (l0 + 1, ligne - 1)  # type: ignore[attr-defined]


def _feuille_scenarios(wb: Workbook, dossier: Dossier) -> None:
    ws_p = wb["Pluie de projet"]
    r0, r1 = ws_p._plage_pluie  # type: ignore[attr-defined]
    ws = wb.create_sheet("Scénarios")
    projet = dossier.projet
    _largeurs(ws, {"A": 52, "B": 16, "C": 16, "D": 16, "E": 16})
    _titre(ws, "A1", "Comparaison des quatre scénarios", 14)
    note = ("Les volumes et durées critiques sont recalculés par formules à partir de la "
            "feuille 'Pluie de projet'.")
    if projet.amont.actif:
        note += (" Ces formules n'appliquent la méthode rationnelle qu'au bassin versant du "
                 "projet : l'apport du bassin d'orage amont varie dans le temps et se poursuit "
                 "après l'averse, il demande une intégration pas à pas qu'une formule de "
                 "cellule ne peut pas reproduire. Le volume à mettre en œuvre est donc celui "
                 "de la ligne « apport du bassin amont compris », reprise de l'application.")
    ws["A2"] = note
    ws["A2"].font = Font(italic=True, color="475569")

    colonnes = {SCENARIO_TEMPORISATION: ("H", "[1]"), SCENARIO_DISPERSION: ("J", "[2]"),
                SCENARIO_MIXTE: ("L", "[3]"), SCENARIO_SEUIL: ("P", "[4]")}
    _entete(ws, 4, ["Grandeur"] + [f"{colonnes[s][1]} {LIBELLES_SCENARIOS[s]}" for s in ORDRE_SCENARIOS])
    ws.row_dimensions[4].height = 46

    def plage(col: str) -> str:
        return f"'Pluie de projet'!${col}${r0}:${col}${r1}"

    lignes: List[Tuple[str, str, str]] = [
        ("Volume à maîtriser [m³]", "=MAX({p})", "0.0"),
        ("Durée critique [min]", "=INDEX('Pluie de projet'!$A${r0}:$A${r1},MATCH(MAX({p}),{p},0))", "0"),
        ("Hauteur de pluie [mm]", "=INDEX('Pluie de projet'!$E${r0}:$E${r1},MATCH(MAX({p}),{p},0))", "0.0"),
        ("Intensité [mm/h]", "=INDEX('Pluie de projet'!$D${r0}:$D${r1},MATCH(MAX({p}),{p},0))", "0.00"),
        ("Intensité [l/s/ha]", "=INDEX('Pluie de projet'!$D${r0}:$D${r1},MATCH(MAX({p}),{p},0))*10000/3600", "0.0"),
        ("Débit ruisselle de pointe [l/s]", "=MAX({p})*0+INDEX('Pluie de projet'!$F${r0}:$F${r1},MATCH(MAX({p}),{p},0))*1000/(INDEX('Pluie de projet'!$A${r0}:$A${r1},MATCH(MAX({p}),{p},0))*60)", "0.00"),
    ]
    ligne = 5
    for libelle, modele, fmt in lignes:
        ws.cell(row=ligne, column=1, value=libelle).font = Font(bold=True)
        for j, s in enumerate(ORDRE_SCENARIOS):
            col = colonnes[s][0]
            c = ws.cell(row=ligne, column=2 + j,
                        value=modele.format(p=plage(col), r0=r0, r1=r1))
            c.number_format = fmt
            c.border = _BORDURE
        ligne += 1

    debits = {
        SCENARIO_TEMPORISATION: "=Q_ajutage",
        SCENARIO_DISPERSION: "=Q_infiltration",
        SCENARIO_MIXTE: "=Q_infiltration+Q_ajutage",
        SCENARIO_SEUIL: "=Q_infiltration+Q_ajutage",
    }
    ws.cell(row=ligne, column=1, value="Débit de sortie [l/s]").font = Font(bold=True)
    for j, s in enumerate(ORDRE_SCENARIOS):
        c = ws.cell(row=ligne, column=2 + j, value=debits[s])
        c.number_format = "0.000"
        c.border = _BORDURE
    l_debit = ligne
    ligne += 1

    ws.cell(row=ligne, column=1, value="Temps de vidange après la pluie [h]").font = Font(bold=True)
    for j, s in enumerate(ORDRE_SCENARIOS):
        col_lettre = get_column_letter(2 + j)
        if s == SCENARIO_SEUIL:
            formule = (f"=IF({col_lettre}{l_debit}<=0,\"\",(MAX({col_lettre}5-V_sous_ajutage,0)*1000/"
                       f"{col_lettre}{l_debit}+MIN({col_lettre}5,V_sous_ajutage)*1000/"
                       f"MAX(Q_infiltration,0.0000001))/3600)")
        else:
            formule = f"=IF({col_lettre}{l_debit}<=0,\"\",{col_lettre}5*1000/{col_lettre}{l_debit}/3600)"
        c = ws.cell(row=ligne, column=2 + j, value=formule)
        c.number_format = "0.0"
        c.border = _BORDURE
        c.fill = PatternFill("solid", fgColor=BLEU_PALE)
    l_vid = ligne
    ligne += 1
    ws.cell(row=ligne, column=1, value="Conforme au temps de vidange maximum ?").font = Font(bold=True)
    for j in range(len(ORDRE_SCENARIOS)):
        col_lettre = get_column_letter(2 + j)
        c = ws.cell(row=ligne, column=2 + j,
                    value=f'=IF({col_lettre}{l_vid}="","-",IF({col_lettre}{l_vid}<=T_vidange_max,"OUI","NON"))')
        c.border = _BORDURE
    ligne += 2

    if projet.amont.actif:
        ws.cell(row=ligne, column=1,
                value="Volume à maîtriser, apport du bassin amont compris [m³]").font = Font(
                    bold=True, color=BLEU)
        for j, s_scen in enumerate(ORDRE_SCENARIOS):
            c = ws.cell(row=ligne, column=2 + j,
                        value=round(dossier.resultats[s_scen].volume_m3, 1))
            c.number_format = "0.0"
            c.border = _BORDURE
            c.fill = PatternFill("solid", fgColor=BLEU_PALE)
        ligne += 2

    ws.cell(row=ligne, column=1, value="Valeurs minimales calculées par l'application").font = Font(bold=True, color=BLEU)
    ligne += 1
    for libelle, cle in (("Surface d'infiltration minimale [m²]", "surface_infiltration_min_m2"),
                         ("Débit d'ajutage minimal [l/s]", "debit_ajutage_min_ls")):
        ws.cell(row=ligne, column=1, value=libelle).font = Font(bold=True)
        for j, s in enumerate(ORDRE_SCENARIOS):
            v = getattr(dossier.resultats[s], cle)
            c = ws.cell(row=ligne, column=2 + j, value="-" if v is None else round(v, 3))
            c.number_format = "0.000"
            c.border = _BORDURE
        ligne += 1

    ligne += 1
    ws.cell(row=ligne, column=1, value="Scénario retenu").font = Font(bold=True)
    ws.cell(row=ligne, column=2, value=LIBELLES_SCENARIOS[dossier.scenario_principal]).font = Font(bold=True, color=BLEU)
    ligne += 2
    alertes = dossier.resultat_principal.alertes + dossier.resultat_principal.messages
    if alertes:
        ws.cell(row=ligne, column=1, value="Observations").font = Font(bold=True, color="B45309")
        ligne += 1
        for a in alertes:
            ws.cell(row=ligne, column=1, value=a).fill = PatternFill("solid", fgColor=ORANGE_PALE)
            ligne += 1


def _feuille_bassin(wb: Workbook, dossier: Dossier) -> None:
    ws = wb.create_sheet("Bassin - table QDF")
    _largeurs(ws, {"A": 22, "B": 14})
    for i in range(len(rainfall.RETURN_PERIODS)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 11
    _titre(ws, "A1", "Bassin encodé - pluies absorbées sans débordement", 14)
    ws["A2"] = ("Volume requis par la pluie (méthode rationnelle) compare à la capacité du bassin. "
                "Vert = absorbe, orange = limite (>95 %), rouge = débordement.")
    ws["A2"].font = Font(italic=True, color="475569")

    projet = dossier.projet
    b = projet.bassin
    _label(ws, 4, "Volume total du bassin", "=V_bassin", "m³", "0.0")
    _label(ws, 5, "Volume sous l'axe de l'ajutage", "=V_sous_ajutage", "m³", "0.0")
    _label(ws, 6, "Surface de dispersion", "=S_dispersion", "m²", "0.0")
    _label(ws, 7, "Débit d'infiltration du bassin", "=Q_infiltration_bassin", "l/s", "0.000")
    _label(ws, 8, "Débit d'ajutage du bassin", "=Q_ajutage_bassin", "l/s", "0.000")

    if dossier.simulation:
        sim = dossier.simulation
        _label(ws, 10, "Événement critique - durée", sim.duree_pluie_min, "min", "0")
        _label(ws, 11, "Événement critique - hauteur", sim.hauteur_pluie_mm, "mm", "0.0")
        _label(ws, 12, "Volume stocké maximum", sim.volume_max_m3, "m³", "0.0")
        _label(ws, 13, "Taux de remplissage", sim.taux_remplissage, "[-]", "0.0%")
        _label(ws, 14, "Volume déborde", sim.volume_debordement_m3, "m³", "0.00")
        _label(ws, 15, "Temps de vidange après la pluie", sim.temps_vidange_h, "h", "0.0")
        _label(ws, 16, "Statut", sim.statut,
               fond=VERT_PALE if not sim.debordement else ROUGE_PALE)

    l0 = 18
    ws.cell(row=l0 - 1, column=1, value="Volume requis [m³] - lignes : durée de pluie, colonnes : période de retour").font = Font(bold=True)
    _entete(ws, l0, ["Durée de pluie"] + [f"{rp} ans" for rp in rainfall.RETURN_PERIODS])
    ws.freeze_panes = f"B{l0 + 1}"

    table = dossier.table
    if table is None:
        return
    for i, duree in enumerate(table.durees_min):
        r = l0 + 1 + i
        ws.cell(row=r, column=1, value=rainfall.QDF_DURATION_LABELS[i] if i < len(rainfall.QDF_DURATION_LABELS)
                else f"{duree:.0f} min").font = Font(bold=True)
        for j, rp in enumerate(table.periodes_retour):
            cell = table.cellules[i][j]
            c = ws.cell(row=r, column=2 + j, value=round(cell.volume_requis_m3, 2))
            c.number_format = "0.0"
            c.border = _BORDURE
            couleur = {"OK": VERT_PALE, "LIMITE": ORANGE_PALE, "DEBORDEMENT": ROUGE_PALE}[cell.statut]
            c.fill = PatternFill("solid", fgColor=couleur)
    l = l0 + 2 + len(table.durees_min)
    rp_max = table.periode_retour_max_acceptee()
    _label(ws, l, "Période de retour maximale absorbée sans débordement",
           f"{rp_max} ans" if rp_max else "aucune (déjà dépassée a 2 ans)",
           fond=VERT_PALE if rp_max else ROUGE_PALE)

    l += 2
    ws.cell(row=l, column=1, value="Hauteurs de pluie correspondantes [mm]").font = Font(bold=True)
    _entete(ws, l + 1, ["Durée de pluie"] + [f"{rp} ans" for rp in rainfall.RETURN_PERIODS])
    for i, duree in enumerate(table.durees_min):
        r = l + 2 + i
        ws.cell(row=r, column=1, value=rainfall.QDF_DURATION_LABELS[i]).font = Font(bold=True)
        for j in range(len(table.periodes_retour)):
            c = ws.cell(row=r, column=2 + j, value=round(table.cellules[i][j].hauteur_mm, 1))
            c.number_format = "0.0"
            c.border = _BORDURE


def _feuille_ajutage(wb: Workbook, dossier: Dossier) -> None:
    ws = wb.create_sheet("Ajutage")
    _largeurs(ws, {"A": 46, "B": 16, "C": 14, "D": 16, "E": 16})
    _titre(ws, "A1", "Dimensionnement de l'ajutage - formule de Torricelli", 14)
    ws["A2"] = "Q = Cd x A x racine(2 g h) - orifice en paroi mince, charge h entre l'axe de l'orifice et le trop-plein."
    ws["A2"].font = Font(italic=True, color="475569")

    _label(ws, 4, "Débit d'ajutage visé", "=IF(Q_ajutage_bassin>0,Q_ajutage_bassin,Q_ajutage)", "l/s", "0.000")
    _label(ws, 5, "Charge h (axe orifice -> trop-plein)", "=Charge_orifice", "m", "0.00")
    _label(ws, 6, "Coefficient de débit Cd", "=Cd_orifice", "[-]", "0.00")
    _label(ws, 7, "Accélération de la pesanteur g", 9.81, "m/s²", "0.00")
    _label(ws, 8, "Section requise A = Q / (Cd.racine(2gh))", "=(B4/1000)/(B6*SQRT(2*B7*B5))", "m²", "0.000000",
           fond=BLEU_PALE)
    _label(ws, 9, "Section requise", "=B8*10000", "cm²", "0.00", fond=BLEU_PALE)
    _label(ws, 10, "Diamètre requis d = racine(4A/pi)", "=2*SQRT(B8/PI())*1000", "mm", "0.0", fond=BLEU_PALE)
    _label(ws, 11, "Vitesse dans l'orifice v = Cd.racine(2gh)", "=B6*SQRT(2*B7*B5)", "m/s", "0.00")

    if dossier.orifice and dossier.orifice.diametre_commercial_mm:
        _label(ws, 12, "Diamètre commercial retenu (par defaut)", dossier.orifice.diametre_commercial_mm, "mm", "0")
        _label(ws, 13, "Débit réel du diamètre retenu",
               "=B6*PI()*(B12/1000)^2/4*SQRT(2*B7*B5)*1000", "l/s", "0.000", fond=VERT_PALE)

    l0 = 16
    ws.cell(row=l0 - 1, column=1, value="Abaque des diamètres commerciaux (charge = h ci-dessus)").font = Font(bold=True)
    _entete(ws, l0, ["Diamètre [mm]", "Section [cm²]", "Débit [l/s]"])
    from ..core.orifice import DIAMETRES_COMMERCIAUX_MM

    for i, d in enumerate(DIAMETRES_COMMERCIAUX_MM):
        r = l0 + 1 + i
        ws.cell(row=r, column=1, value=d).border = _BORDURE
        ws.cell(row=r, column=2, value=f"=PI()*(A{r}/1000)^2/4*10000").number_format = "0.00"
        ws.cell(row=r, column=3, value=f"=$B$6*PI()*(A{r}/1000)^2/4*SQRT(2*$B$7*$B$5)*1000").number_format = "0.000"


def _feuille_statistiques(wb: Workbook, dossier: Dossier) -> None:
    projet = dossier.projet
    ws = wb.create_sheet("Pluies statistiques")
    _largeurs(ws, {"A": 16})
    for i in range(len(rainfall.RETURN_PERIODS)):
        ws.column_dimensions[get_column_letter(2 + i)].width = 10
    _titre(ws, "A1", f"Pluies statistiques GTI - {projet.commune_nom} ({projet.commune_ins})", 14)
    ws["A2"] = dossier.libelle_source
    ws["A2"].font = Font(italic=True, color="475569")

    mm = rainfall.table_qdf_mm(projet.commune_ins, projet.source_pluie)
    lsha = rainfall.table_qdf_ls_ha(projet.commune_ins, projet.source_pluie)
    for titre, table, fmt, depart in (("Hauteurs de pluie [mm]", mm, "0.0", 4),
                                      ("Intensités [l/s/ha]", lsha, "0.0", 4 + len(mm) + 4)):
        ws.cell(row=depart - 1, column=1, value=titre).font = Font(bold=True, color=BLEU)
        _entete(ws, depart, ["Durée"] + [f"{rp} ans" for rp in rainfall.RETURN_PERIODS])
        for i, ligne in enumerate(table):
            r = depart + 1 + i
            ws.cell(row=r, column=1, value=rainfall.QDF_DURATION_LABELS[i]).font = Font(bold=True)
            for j, v in enumerate(ligne):
                c = ws.cell(row=r, column=2 + j, value=None if v is None else round(v, 2))
                c.number_format = fmt
                c.border = _BORDURE
                if rainfall.RETURN_PERIODS[j] == projet.periode_retour:
                    c.fill = PatternFill("solid", fgColor=BLEU_PALE)


def ecrire(dossier: Dossier, chemin: str) -> str:
    """Génère le rapport Excel et renvoie le chemin du fichier."""
    wb = construire_classeur(dossier)
    wb.save(chemin)
    return chemin
